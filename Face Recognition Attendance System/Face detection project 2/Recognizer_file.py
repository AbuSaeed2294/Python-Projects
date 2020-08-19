import cv2
import openpyxl

face_cas = cv2.CascadeClassifier('D:/Programs/Python Project/FINAL YEAR PROJECT 2/Face detection project 2'
                                 '/haarcascade_frontalface_default.xml')
cap = cv2.VideoCapture(0)
cap.set(3, 1080)
cap.set(4, 720)

recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read('D:/Programs/Python Project/FINAL YEAR PROJECT 2/Face detection project 2/trainer.yml')

font = cv2.FONT_HERSHEY_SIMPLEX


while True:

    ret, img = cap.read()
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = face_cas.detectMultiScale(gray, 1.1, 5)

    for (x,y,w,h) in faces:
        roi_gray = gray[y:y + h, x:x + w]
        cv2.rectangle(img, (x,y), (x+w, y+h), (255,0,0),2)
        id,conf=recognizer.predict(roi_gray)

        if conf < 50:
            id = id
            if id == 646:
                my_wb = openpyxl.Workbook()
                my_sheet = my_wb.active
                c1 = my_sheet.cell(row=2, column=1)
                c1.value = id
                c2 = my_sheet.cell(row=2, column=2)
                c2.value = 'Abu saeed'
                c3 = my_sheet.cell(row=2, column=3)
                c3.value = 'Present'
                my_wb.save("D://Programs//Python Project//FINAL YEAR PROJECT 2//Attendance sheet//"
                           "Abu saeed.xlsx")

            elif id == 123:
                my_wb = openpyxl.Workbook()
                my_sheet = my_wb.active
                c1 = my_sheet.cell(row=3, column=1)
                c1.value = id
                c2 = my_sheet.cell(row=3, column=2)
                c2.value = 'Guest'
                c3 = my_sheet.cell(row=3, column=3)
                c3.value = 'Present'
                my_wb.save("D://Programs//Python Project//FINAL YEAR PROJECT 2//Attendance sheet//Guest.xlsx")

            elif id == 1234:
                my_wb = openpyxl.Workbook()
                my_sheet = my_wb.active
                c1 = my_sheet.cell(row=3, column=1)
                c1.value = id
                c2 = my_sheet.cell(row=3, column=2)
                c2.value = 'Cecos'
                c3 = my_sheet.cell(row=3, column=3)
                c3.value = 'Present'
                my_wb.save("D://Programs//Python Project//FINAL YEAR PROJECT 2//Attendance sheet//Cecos.xlsx")

            conf = '{0}%'. format(round(100 - conf))

        elif conf > 50:
            id = 'Unknown'
            conf = '{0}%'.format(round(100 - conf))

        else:
            pass

        cv2.putText(img, str(id), (x + 5, x - 5), font, 1, (255, 255, 255), 2)
        cv2.putText(img, str(conf), (x + 5, y + h - 5), font, 1, (255, 255, 0), 1)

    cv2.imshow('Camera', img)

    k = cv2.waitKey(10) & 0xff
    if k == 27:
        break


cap.release()
cv2.destroyAllWindows()